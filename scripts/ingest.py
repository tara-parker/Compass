#!/usr/bin/env python3
"""
Compass ingest
==============
Turns Google Search Console + keyword-rank exports into a nested markdown
content tree that Compass reads. Re-run any time you add a new weekly export;
each run rebuilds the tree and appends the new period as another snapshot on
every page, cluster and the site pillar.

Inputs  (data/source/):
  - "*Performance-on-Search*.xlsx"   one file per GSC weekly export
  - "*keyword*rank*.xlsx"            keyword position tracker (optional)

Outputs:
  - content/_root.md                 main pillar (whole-site totals over time)
  - content/<cluster>/_cluster.md    section aggregate over time
  - content/<cluster>/<slug>.md      individual tracked page over time
  - data/keywords.json               keyword rankings for the /keywords view
  - data/meta.json                   period list + generated-at

Usage:  python3 scripts/ingest.py
"""

import glob
import json
import os
import re
import datetime
from collections import defaultdict, Counter
from urllib.parse import urlparse

try:
    import openpyxl
except ImportError:
    raise SystemExit("openpyxl is required:  pip3 install openpyxl")

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(ROOT, "data", "source")
CONTENT = os.path.join(ROOT, "content")
DATA = os.path.join(ROOT, "data")

MONTHS = {m: i for i, m in enumerate(
    ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
     "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"], start=1)}


# --------------------------------------------------------------------------- #
# helpers
# --------------------------------------------------------------------------- #
def sheet_rows(ws):
    for r in ws.iter_rows(values_only=True):
        if any(c is not None for c in r):
            yield r


def parse_gsc_date_range(label):
    """'Aug 3, 2026-Aug 10, 2026' -> (start_iso, end_iso, human)."""
    m = re.findall(r"([A-Za-z]{3}) (\d{1,2}), (\d{4})", label or "")
    if len(m) != 2:
        return (label, label, label)
    (m1, d1, y1), (m2, d2, y2) = m
    start = datetime.date(int(y1), MONTHS[m1], int(d1))
    end = datetime.date(int(y2), MONTHS[m2], int(d2))
    human = f"{m1} {int(d1)}–{m2} {int(d2)}" if m1 != m2 else f"{m1} {int(d1)}–{int(d2)}"
    return (start.isoformat(), end.isoformat(), human)


def num(v):
    if v is None:
        return None
    try:
        return round(float(v), 4)
    except (TypeError, ValueError):
        return None


def w_avg_pos(items):
    """Impression-weighted average position (the GSC-correct aggregate)."""
    tot_impr = sum((i["impressions"] or 0) for i in items)
    if not tot_impr:
        vals = [i["position"] for i in items if i["position"] is not None]
        return round(sum(vals) / len(vals), 2) if vals else None
    s = sum((i["position"] or 0) * (i["impressions"] or 0) for i in items)
    return round(s / tot_impr, 2)


def agg(items):
    clicks = sum((i["clicks"] or 0) for i in items)
    impr = sum((i["impressions"] or 0) for i in items)
    return {
        "clicks": int(clicks),
        "impressions": int(impr),
        "ctr": round(clicks / impr, 4) if impr else 0.0,
        "position": w_avg_pos(items),
    }


ACRONYMS = {
    "ai": "AI", "erp": "ERP", "cfo": "CFO", "cfos": "CFOs", "ceo": "CEO",
    "coo": "COO", "cto": "CTO", "api": "API", "apis": "APIs", "sap": "SAP",
    "ap": "AP", "ar": "AR", "roi": "ROI", "kpi": "KPI", "kpis": "KPIs",
    "saas": "SaaS", "b2b": "B2B", "us": "US", "uk": "UK", "hr": "HR",
    "it": "IT", "ml": "ML", "nlp": "NLP", "crm": "CRM", "fpa": "FP&A",
    "ebs": "EBS", "sql": "SQL", "gpt": "GPT", "llm": "LLM", "llms": "LLMs",
    "vs": "vs", "netsuite": "NetSuite", "quickbooks": "QuickBooks",
    "floqast": "FloQast", "blackline": "BlackLine", "chatfin": "ChatFin",
    "d365": "D365",
}


def title_from_slug(slug):
    s = slug.replace("-", " ").replace("_", " ").strip()
    small = {"a", "an", "and", "the", "of", "for", "to", "in", "on", "with", "at"}
    words = []
    for i, w in enumerate(s.split()):
        lw = w.lower()
        if lw in ACRONYMS:
            words.append(ACRONYMS[lw])
        elif w.isupper() or any(c.isdigit() for c in w):
            words.append(w)
        elif i > 0 and lw in small:
            words.append(lw)
        else:
            words.append(w.capitalize())
    return " ".join(words) or slug


def cluster_title(name):
    override = {
        "ai-erp": "AI ERP", "fpa-ai-insights": "FP&A AI Insights",
        "wp-content": "WP Content",
    }
    if name in override:
        return override[name]
    return title_from_slug(name)


# --------------------------------------------------------------------------- #
# YAML emitter (tiny, tailored to the shapes we write)
# --------------------------------------------------------------------------- #
def yq(v):
    if v is None:
        return "null"
    if isinstance(v, bool):
        return "true" if v else "false"
    if isinstance(v, (int, float)):
        return repr(v)
    s = str(v)
    # Quote ISO dates / date-like tokens so YAML parsers keep them as strings.
    if re.match(r"^\d{4}-\d{2}-\d{2}", s):
        return '"' + s + '"'
    if s == "" or re.search(r'[:#\[\]{}\",\']', s) or s != s.strip():
        return '"' + s.replace("\\", "\\\\").replace('"', '\\"') + '"'
    return s


def emit_snapshot_list(key, snaps):
    lines = [f"{key}:"]
    if not snaps:
        return [f"{key}: []"]
    for s in snaps:
        lines.append(f"  - period: {yq(s['period'])}")
        lines.append(f"    label: {yq(s['label'])}")
        lines.append(f"    start: {yq(s['start'])}")
        lines.append(f"    end: {yq(s['end'])}")
        lines.append(f"    clicks: {yq(s['clicks'])}")
        lines.append(f"    impressions: {yq(s['impressions'])}")
        lines.append(f"    ctr: {yq(s['ctr'])}")
        lines.append(f"    position: {yq(s['position'])}")
    return lines


def emit_daily(daily):
    lines = ["daily:"]
    if not daily:
        return ["daily: []"]
    for d in daily:
        lines.append(
            f"  - {{ date: {yq(d['date'])}, clicks: {yq(d['clicks'])}, "
            f"impressions: {yq(d['impressions'])}, ctr: {yq(d['ctr'])}, "
            f"position: {yq(d['position'])} }}"
        )
    return lines


def write_md(path, front_lines, body):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        f.write("---\n")
        f.write("\n".join(front_lines))
        f.write("\n---\n\n")
        f.write(body.rstrip() + "\n")


# --------------------------------------------------------------------------- #
# load GSC exports
# --------------------------------------------------------------------------- #
def load_gsc():
    files = glob.glob(os.path.join(SRC, "*Performance-on-Search*.xlsx"))
    files += glob.glob(os.path.join(SRC, "*performance*.xlsx"))
    files = sorted(set(files))
    periods = []
    for fp in files:
        wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
        filt = {}
        if "Filters" in wb.sheetnames:
            for r in sheet_rows(wb["Filters"]):
                if r and r[0]:
                    filt[str(r[0])] = r[1]
        start, end, human = parse_gsc_date_range(str(filt.get("Date", os.path.basename(fp))))
        pid = f"{start}_{end}"

        pages = {}
        if "Pages" in wb.sheetnames:
            rows = list(sheet_rows(wb["Pages"]))[1:]
            for r in rows:
                url = r[0]
                if not url or "chatfin.ai" not in str(url):
                    continue
                pages[str(url)] = {
                    "clicks": num(r[1]) or 0,
                    "impressions": num(r[2]) or 0,
                    "ctr": num(r[3]) or 0,
                    "position": num(r[4]),
                }

        daily = []
        if "Chart" in wb.sheetnames:
            for r in list(sheet_rows(wb["Chart"]))[1:]:
                d = r[0]
                if isinstance(d, (datetime.datetime, datetime.date)):
                    d = d.isoformat()[:10]
                daily.append({
                    "date": str(d), "clicks": int(num(r[1]) or 0),
                    "impressions": int(num(r[2]) or 0),
                    "ctr": num(r[3]) or 0, "position": num(r[4]),
                })

        queries = []
        if "Queries" in wb.sheetnames:
            for r in list(sheet_rows(wb["Queries"]))[1:]:
                if not r or not r[0]:
                    continue
                queries.append({
                    "query": str(r[0]), "clicks": int(num(r[1]) or 0),
                    "impressions": int(num(r[2]) or 0),
                    "ctr": num(r[3]) or 0, "position": num(r[4]),
                })

        wb.close()
        periods.append({
            "id": pid, "label": human, "start": start, "end": end,
            "file": os.path.basename(fp), "pages": pages, "daily": daily,
            "queries": queries,
        })
    periods.sort(key=lambda p: p["start"])
    return periods


# --------------------------------------------------------------------------- #
# hierarchy
# --------------------------------------------------------------------------- #
def url_to_node(url):
    """Return (cluster, rel_path_parts, is_home)."""
    path = urlparse(url).path.strip("/")
    if path == "":
        return (None, [], True)
    parts = [p for p in path.split("/") if p]
    cluster = parts[0]
    rest = parts[1:] if len(parts) > 1 else [parts[0]]
    # a top-level page like /about -> cluster 'pages', slug 'about'
    if len(parts) == 1:
        return ("pages", [parts[0]], False)
    return (cluster, rest, False)


def snap_for(period, stats):
    return {
        "period": period["id"], "label": period["label"],
        "start": period["start"], "end": period["end"],
        "clicks": int(stats["clicks"]), "impressions": int(stats["impressions"]),
        "ctr": round(stats["ctr"], 4), "position": stats["position"],
    }


def main():
    periods = load_gsc()
    if not periods:
        raise SystemExit("No GSC exports found in data/source/")

    print(f"Loaded {len(periods)} period(s): " +
          ", ".join(p["label"] for p in periods))

    # wipe generated content (keep nothing stale), then rebuild
    if os.path.isdir(CONTENT):
        for dirpath, _dirs, files in os.walk(CONTENT):
            for fn in files:
                if fn.endswith(".md"):
                    os.remove(os.path.join(dirpath, fn))

    # --- collect every url across all periods ---
    all_urls = set()
    for p in periods:
        all_urls.update(p["pages"].keys())

    # group urls -> (cluster, slug filename, rel parts)
    page_index = {}          # url -> dict(cluster, parts, slug, title)
    cluster_pages = defaultdict(list)   # cluster -> [url]
    home_url = None
    for url in all_urls:
        cluster, parts, is_home = url_to_node(url)
        if is_home:
            home_url = url
            continue
        page_index[url] = {
            "cluster": cluster, "parts": parts,
            "slug": "/".join(parts),
            "title": title_from_slug(parts[-1]),
            "url": url,
        }
        cluster_pages[cluster].append(url)

    # fold this run into the persistent update ledger (survives content wipe)
    hist = record_updates(page_index, periods, home_url)

    # ----------------------------------------------------------------- pages
    n_pages = 0
    for url, info in page_index.items():
        snaps = []
        for p in periods:
            if url in p["pages"]:
                snaps.append(snap_for(p, p["pages"][url]))
        latest = snaps[-1] if snaps else None
        prev = snaps[-2] if len(snaps) > 1 else None

        front = [
            f"title: {yq(info['title'])}",
            f"url: {yq(url)}",
            "type: sub",
            f"cluster: {yq(info['cluster'])}",
            f"slug: {yq(info['slug'])}",
        ]
        front += emit_snapshot_list("snapshots", snaps)
        front += emit_updates_front(hist["pages"].get(url))

        body = build_page_body(info, snaps, latest, prev)
        rel = os.path.join(info["cluster"], *info["parts"]) + ".md"
        write_md(os.path.join(CONTENT, rel), front, body)
        n_pages += 1

    # -------------------------------------------------------------- clusters
    for cluster, urls in cluster_pages.items():
        snaps = []
        for p in periods:
            items = [p["pages"][u] for u in urls if u in p["pages"]]
            if items:
                snaps.append(snap_for(p, agg(items)))
        # page count present in latest period
        latest_period = periods[-1]
        pages_live = sum(1 for u in urls if u in latest_period["pages"])
        front = [
            f"title: {yq(cluster_title(cluster))}",
            "type: cluster",
            f"cluster: {yq(cluster)}",
            f"pages_tracked: {len(urls)}",
            f"pages_live: {pages_live}",
        ]
        front += emit_snapshot_list("snapshots", snaps)
        body = build_cluster_body(cluster, urls, snaps)
        write_md(os.path.join(CONTENT, cluster, "_cluster.md"), front, body)

    # ------------------------------------------------------------------ root
    root_snaps = []
    daily_map = {}
    for p in periods:
        items = list(p["pages"].values())
        root_snaps.append(snap_for(p, agg(items)))
        for d in p["daily"]:
            daily_map[d["date"]] = d   # later export wins on overlapping dates
    daily_all = [daily_map[k] for k in sorted(daily_map)]
    home_snaps = []
    if home_url:
        for p in periods:
            if home_url in p["pages"]:
                home_snaps.append(snap_for(p, p["pages"][home_url]))

    front = [
        "title: ChatFin — Site Overview",
        "type: main",
        f"url: {yq(home_url or 'https://chatfin.ai/')}",
        f"clusters: {len(cluster_pages)}",
        f"pages_tracked: {len(page_index)}",
    ]
    front += emit_snapshot_list("snapshots", root_snaps)
    front += emit_snapshot_list("homepage", home_snaps)
    front += emit_daily(daily_all)
    body = build_root_body(root_snaps, cluster_pages, periods)
    write_md(os.path.join(CONTENT, "_root.md"), front, body)

    # --------------------------------------------------------------- updates
    ufront, ubody = build_updates_page(hist)
    write_md(os.path.join(CONTENT, "_updates.md"), ufront, ubody)

    # -------------------------------------------------------------- keywords
    keywords = load_keywords()
    with open(os.path.join(DATA, "keywords.json"), "w", encoding="utf-8") as f:
        json.dump(keywords, f, indent=2)

    # ------------------------------------------------------------------ meta
    meta = {
        "periods": [{"id": p["id"], "label": p["label"],
                     "start": p["start"], "end": p["end"]} for p in periods],
        "pages_tracked": len(page_index),
        "clusters": len(cluster_pages),
        "generated_note": "Run `npm run ingest` after adding a new export.",
    }
    with open(os.path.join(DATA, "meta.json"), "w", encoding="utf-8") as f:
        json.dump(meta, f, indent=2)

    n_events = sum(len(r.get("events", [])) for r in hist["pages"].values())
    print(f"Wrote {n_pages} page files across {len(cluster_pages)} clusters.")
    print(f"Update log: {n_events} events across {len(hist['periods'])} dates.")
    print(f"Keywords: {len(keywords.get('summary', []))} summarized, "
          f"{len(keywords.get('tracked', []))} tracked.")


# --------------------------------------------------------------------------- #
# markdown bodies
# --------------------------------------------------------------------------- #
def _delta_word(latest, prev, field, lower_better=False):
    if not latest or not prev:
        return ""
    a, b = prev.get(field), latest.get(field)
    if a is None or b is None:
        return ""
    diff = b - a
    if abs(diff) < 1e-9:
        return "held steady"
    if lower_better:
        return "improved" if diff < 0 else "slipped"
    return "rose" if diff > 0 else "fell"


def build_page_body(info, snaps, latest, prev):
    # Body is the editable notes area only. All metrics live in frontmatter and
    # are rendered by the UI, so we deliberately do not repeat them here.
    return "_No notes yet. Edit this file to add observations for this page._"


def build_cluster_body(cluster, urls, snaps):
    return f"_Editable notes for the {cluster_title(cluster)} cluster._"


def build_root_body(root_snaps, cluster_pages, periods):
    return "_Editable notes for the site pillar._"


def fmt(v):
    return "—" if v is None else f"{v:.1f}"


def pct(v):
    return "—" if v is None else f"{v * 100:.2f}%"


# --------------------------------------------------------------------------- #
# keywords
# --------------------------------------------------------------------------- #
def load_keywords():
    files = glob.glob(os.path.join(SRC, "*keyword*rank*.xlsx"))
    files += glob.glob(os.path.join(SRC, "*rankings*.xlsx"))
    files = sorted(set(files))
    if not files:
        return {"summary": [], "tracked": [], "ahrefs": []}
    wb = openpyxl.load_workbook(files[0], read_only=True, data_only=True)
    out = {"summary": [], "tracked": [], "ahrefs": [], "source": os.path.basename(files[0])}

    if "main" in wb.sheetnames:
        rows = list(sheet_rows(wb["main"]))[1:]
        for r in rows:
            out["tracked"].append({
                "keyword": str(r[1]) if len(r) > 1 and r[1] is not None else "",
                "url": str(r[2]) if len(r) > 2 and r[2] is not None else "",
                "position": num(r[3]) if len(r) > 3 else None,
                "variation": str(r[4]) if len(r) > 4 and r[4] is not None else "",
            })

    if "summary" in wb.sheetnames:
        rows = list(sheet_rows(wb["summary"]))[1:]
        for r in rows:
            out["summary"].append({
                "keyword": str(r[0]) if r[0] is not None else "",
                "best_position": num(r[1]) if len(r) > 1 else None,
                "pages_ranking": num(r[2]) if len(r) > 2 else None,
                "positions": str(r[3]) if len(r) > 3 and r[3] is not None else "",
                "variation": str(r[4]) if len(r) > 4 and r[4] is not None else "",
            })

    if "ahrefs" in wb.sheetnames:
        rows = list(sheet_rows(wb["ahrefs"]))[1:]
        for r in rows:
            out["ahrefs"].append({
                "keyword": str(r[0]) if r[0] is not None else "",
                "skill_pos": str(r[1]) if len(r) > 1 and r[1] is not None else "",
                "ahrefs_pos": str(r[2]) if len(r) > 2 and r[2] is not None else "",
                "verdict": str(r[3]) if len(r) > 3 and r[3] is not None else "",
            })

    wb.close()
    return out




# ---------------------------------------------------------------- update log
# The content/ tree is wiped and regenerated on every ingest, so update history
# lives in a persistent JSON ledger under data/. Each ingest appends new events
# and never rewrites past ones, which makes the log durable across rebuilds.

HISTORY_PATH = os.path.join(ROOT, "data", "history.json")

# Fields we treat as meaningful movement for a page.
TRACKED_FIELDS = ("position", "clicks", "impressions")


def load_history():
    try:
        with open(HISTORY_PATH, "r", encoding="utf-8") as fh:
            data = json.load(fh)
    except (FileNotFoundError, ValueError):
        return {"pages": {}, "periods": []}
    data.setdefault("pages", {})
    data.setdefault("periods", [])
    return data


def save_history(hist):
    os.makedirs(os.path.dirname(HISTORY_PATH), exist_ok=True)
    with open(HISTORY_PATH, "w", encoding="utf-8") as fh:
        json.dump(hist, fh, indent=2, sort_keys=True)
        fh.write("\n")


def _round(v, nd=2):
    return None if v is None else round(float(v), nd)


def classify(prev, cur):
    """Describe what changed between two consecutive snapshots of one page."""
    if prev is None:
        return "added", "first tracked"
    def i(v):
        return int(v or 0)

    bits = []
    dp = (cur["position"] or 0) - (prev["position"] or 0)
    if abs(dp) >= 0.05:
        bits.append(("improved" if dp < 0 else "declined",
                     f"position {_round(prev['position'], 1)} to {_round(cur['position'], 1)}"))
    if i(cur["clicks"]) != i(prev["clicks"]):
        bits.append((None, f"clicks {i(prev['clicks'])} to {i(cur['clicks'])}"))
    if i(cur["impressions"]) != i(prev["impressions"]):
        bits.append((None, f"impressions {i(prev['impressions'])} to {i(cur['impressions'])}"))
    if not bits:
        return None, None
    # Position is the headline signal for a rank tracker; fall back to "changed"
    # when only click/impression volume moved.
    kind = next((k for k, _ in bits if k), "changed")
    return kind, "; ".join(t for _, t in bits)


def record_updates(page_index, periods, home_url):
    """Fold this run's periods into the persistent ledger, returning it."""
    hist = load_history()
    seen_periods = {p["id"] for p in hist["periods"]}
    for p in periods:
        if p["id"] not in seen_periods:
            hist["periods"].append({"id": p["id"], "label": p["label"],
                                    "start": p["start"], "end": p["end"]})
    hist["periods"].sort(key=lambda x: x["end"])

    targets = dict(page_index)
    if home_url:
        targets[home_url] = {"slug": "", "cluster": "", "title": "ChatFin — Site Overview"}

    for url, info in targets.items():
        rec = hist["pages"].setdefault(url, {"events": [], "first_seen": None})
        logged = {e["period"] for e in rec["events"]}
        prev = None
        for p in periods:
            stats = p["pages"].get(url)
            if not stats:
                continue
            cur = {"position": num(stats.get("position")),
                   "clicks": num(stats.get("clicks")),
                   "impressions": num(stats.get("impressions"))}
            if rec["first_seen"] is None:
                rec["first_seen"] = p["end"]
            if p["id"] not in logged:
                kind, detail = classify(prev, cur)
                if kind:
                    rec["events"].append({
                        "date": p["end"], "period": p["id"], "label": p["label"],
                        "kind": kind, "detail": detail,
                        "position": _round(cur["position"], 2),
                        "clicks": int(cur["clicks"] or 0),
                        "impressions": int(cur["impressions"] or 0),
                    })
            prev = cur
        rec["events"].sort(key=lambda e: e["date"])
        rec["slug"] = info.get("slug", "")
        rec["cluster"] = info.get("cluster", "")
        rec["title"] = info.get("title", "")

    save_history(hist)
    return hist


def emit_updates_front(rec):
    """Frontmatter lines describing a page's update history."""
    if not rec:
        return []
    events = rec.get("events", [])
    out = []
    if rec.get("first_seen"):
        out.append(f"first_seen: {yq(rec['first_seen'])}")
    if events:
        out.append(f"last_updated: {yq(events[-1]['date'])}")
        out.append(f"update_count: {len(events)}")
        out.append("updates:")
        for e in events:
            out.append(f"  - date: {yq(e['date'])}")
            out.append(f"    period: {yq(e['period'])}")
            out.append(f"    kind: {yq(e['kind'])}")
            out.append(f"    detail: {yq(e['detail'])}")
    return out


def build_updates_page(hist, top_n=40):
    """Timeline digest: what changed on each date, plus that date's top movers.

    Every page keeps its own full history in its own frontmatter, so this file
    stays a readable digest rather than a dump of every event.
    """
    by_date = defaultdict(list)
    for url, rec in hist["pages"].items():
        for e in rec.get("events", []):
            by_date[e["date"]].append((rec.get("slug") or "(home)", rec.get("title") or url,
                                       rec.get("cluster", ""), e))

    def impact(row):
        e = row[3]
        return (e.get("clicks") or 0, e.get("impressions") or 0)

    front = ["title: Update log", "type: updates",
             f"dates_tracked: {len(by_date)}",
             f"pages_tracked: {len(hist['pages'])}",
             f"top_n: {top_n}", "log:"]
    for date in sorted(by_date, reverse=True):
        rows = by_date[date]
        counts = Counter(e["kind"] for _, _, _, e in rows)
        front.append(f"  - date: {yq(date)}")
        front.append(f"    pages_changed: {len(rows)}")
        front.append(f"    added: {counts.get('added', 0)}")
        front.append(f"    improved: {counts.get('improved', 0)}")
        front.append(f"    declined: {counts.get('declined', 0)}")
        front.append(f"    changed: {counts.get('changed', 0)}")
        front.append("    movers:")
        for slug, title, cluster, e in sorted(rows, key=impact, reverse=True)[:top_n]:
            front.append(f"      - slug: {yq(slug)}")
            front.append(f"        title: {yq(title)}")
            front.append(f"        cluster: {yq(cluster)}")
            front.append(f"        kind: {yq(e['kind'])}")
            front.append(f"        clicks: {yq(e.get('clicks'))}")
            front.append(f"        position: {yq(e.get('position'))}")
            front.append(f"        detail: {yq(e['detail'])}")

    body = ("_Which tracked page moved, and when. Dates are the close of each "
            "Search Console window, so an entry means the page's tracked "
            "metrics changed as of that date._")
    return front, body


if __name__ == "__main__":
    main()
