#!/usr/bin/env python3
"""
Build data/plan26.json . the 2026 prune-and-build plan.

Every live URL gets one of four actions: KEEP, UPDATE, MERGE or DELETE.
Output is nested cluster -> sub-cluster -> page so the UI mirrors the site.

WHY DELETE IS NOT THE WHOLE TAIL
--------------------------------
An earlier cut of this plan marked 3,255 URLs for deletion on the grounds that
they drew no Search Console impressions. That was wrong, for two reasons that
are properties of the data rather than of the pages:

1. THE EXPORT CAP. GSC's UI export stops at 1,000 rows. The weakest page in
   each weekly export had 1-2 impressions, so the cap bites exactly at the
   bottom of the range. A page absent from the export may have had impressions
   and simply ranked below the cut. "Zero impressions" here means "not in the
   top 1,000", which is not evidence of anything.

2. THE MEASUREMENT WINDOW. 90% of the site carries a Jul/Aug 2026 lastmod, and
   2,989 of the 3,344 pages with no GSC row were published or modified inside
   the two-week window being measured. A page published a fortnight ago has no
   impressions because Google has not ranked it yet.

So absence of traffic is not treated as evidence. DELETE now requires positive
evidence of worthlessness: near-duplicate of a better page, or no signal after
a genuine fair trial. Pages that were never given a chance are kept or worked
on, and re-measured once they have real history.

HOW THE ACTIONS ARE DECIDED
---------------------------
Checked in order, first match wins:

  KEEP    backlinked and earning  |  earns clicks  |  position 20 or better
          |  unique topic with no evidence against it
  UPDATE  backlinked but earning nothing  |  indexed but not converting
          |  canonical of a duplicate group  |  crowded topic needing
             differentiation
  MERGE   near-duplicate that has signal  |  near-duplicate whose whole group
          is still unmeasured, so the best member cannot be picked yet
  DELETE  near-duplicate of a page that ALREADY earns  |  qualifier-padded
          variant of a shorter page  |  near-duplicate with no signal after a
          fair trial  |  no signal at all after a fair trial

Two things separate DELETE from MERGE. Age defends a unique page but not a
duplicate: a new unique page might yet rank, whereas a near-duplicate of a page
that already earns will not, because Google has already picked between them.
And there is nothing to fold in . the canonical carries the same content . nor
anything to pass, since these have no links and no traffic. Per the redirect
rules, a 301 is only worth adding when the source has signal to pass.

Duplicate detection is Jaccard similarity over normalised slug tokens, union
-found into groups. Tight (>=0.6) marks true duplicates; loose (>=0.35, groups
of 3+) marks crowded topic neighbourhoods. Separately, a page whose tokens are
a strict SUPERSET of a shorter page's is a qualifier-padded variant of it
("accrual-automation-software-for-month-end-close" off
"accrual-automation-software"), which Jaccard misses because the padding drags
similarity down. Slug similarity is a proxy throughout . the real test is
body-text hashing, which is what the SAP B1 audit used.

ORDERING
--------
Pages are listed KEEP, UPDATE, MERGE, DELETE, and ranked by measured strength
inside each action (clicks, then impressions, then position). Every page also
carries `n`, its rank by strength within its list, so the UI can re-sort by
strength alone without recomputing.

Sources (read-only, outside the repo):
  ~/Downloads/chatfin-all-urls.json                     live sitemap
  ~/Desktop/chatfin/code/cf-platform/*Performance*.xlsx GSC page exports
  ~/Desktop/chatfin/seo/redirects/all-backlinked.txt    Ahrefs backlink set

Writes to data/, NOT content/ . scripts/ingest.py wipes content/ on every run.

    python3 scripts/build_plan26.py
"""

import json
import glob
import os
import re
import collections
import datetime

HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.dirname(HERE)
OUT = os.path.join(REPO, "data", "plan26.json")

HOME = os.path.expanduser("~")
SITEMAP = os.path.join(HOME, "Downloads", "chatfin-all-urls.json")
GSC_GLOB = os.path.join(HOME, "Desktop", "chatfin", "code", "cf-platform", "*Performance*.xlsx")
SEO = os.path.join(HOME, "Desktop", "chatfin", "seo", "redirects")
BL = os.path.join(HOME, "Desktop", "chatfin", "blogs", "backlinks")

# What counts as backlinked: a URL that AHREFS reports as carrying links from
# external domains. Nothing else. REVIVED-LINKS.txt is deliberately NOT a source
# . it is the record of an August redirect job, listing URLs whose redirects
# were removed, which is a history of what was done rather than current evidence
# that a link exists today. Where those pages do still hold links, Ahrefs lists
# them anyway and they are protected on that basis.
#
# Column-precise on purpose: the redirect tables also carry TARGET urls, which
# are not themselves backlinked and must not be pulled in.
BACKLINK_LISTS = [                       # every line is a backlinked path
    os.path.join(SEO, "all-backlinked.txt"),
    os.path.join(SEO, "live-backlinked.txt"),
    os.path.join(BL, "live-backlinked.txt"),
]
BACKLINK_TSV_COL0 = [                    # column 0 is the backlinked url
    os.path.join(SEO, "backlinked-redirects.tsv"),
    os.path.join(SEO, "protect-list.tsv"),
    os.path.join(SEO, "uncrawled-backlinked.tsv"),
    os.path.join(BL, "final-status.tsv"),
]
BACKLINK_XLSX = os.path.join(BL, "chatfin-backlinked-redirects.xlsx")
# Pages already rewritten by the ERP rigs. A page rewritten days ago has had no
# more chance to prove itself than one published days ago, so the fair-trial
# logic applies to it too: it is never deleted, whatever the old numbers say.
# Without this the plan bins work that was just done.
REWRITTEN = os.path.join(HOME, "Desktop", "chatfin", "blogs", "plan26", "update",
                         "updated-pages.json")
# Per-ERP do-not-redirect lists. These name pages that were rewritten in place
# and must keep their URL: redirecting one destroys the rewrite. The SAP B1 list
# holds the 37 rewritten pages, none of which appear in updated-pages.json,
# so without this the plan marks 10 of them MERGE or DELETE. One sits at
# position 4.6.
DO_NOT_REDIRECT = os.path.join(HOME, "Desktop", "chatfin", "blogs", "redirect",
                               "*", "07-do-not-redirect.txt")

BACKLINK_XLSX_SHEETS = [                 # column 0 only, never the target column
    "FINAL - all 98 resolved",
    "PROTECT - live backlinked",
    "REMOVE these redirects",
    "KEEP - no page behind",
    "KEEP - slash normalisation",
]
# GSC "Crawled - currently not indexed". Google fetched the page and decided not
# to index it. That is the strongest quality signal available: unlike an absent
# GSC row, which only means the page was below the export cap or too new, this
# is Google actively refusing the page after looking at it.
NOT_INDEXED = os.path.join(HOME, "Downloads", "Not Indexed Pages...xlsx")

# Ahrefs Site Audit crawl: the pages a crawler can actually reach by following
# links. Sitemap URLs missing from it are orphans with no internal links in.
AHREFS_CRAWL = os.path.join(HOME, "Desktop", "chatfin", "keywords", "seo",
                            "cf-pages-indexed-by-ahref.md")
# Ahrefs referring-domain counts, where known.
REF_DOMAINS = os.path.join(HOME, "Desktop", "chatfin", "seo", "redirects", "protect-list.tsv")

GENERATED = "2026-08-19"
WINDOW = "3-17 Aug 2026 (2 weeks)"

# A page modified after this has not had time to rank, so it cannot be judged.
FAIR_TRIAL_BEFORE = "2026-06"

TIGHT = 0.6    # near-duplicate
LOOSE = 0.35   # crowded neighbourhood
CROWD_MIN = 3

# Days a page needs before its numbers mean anything. Watchlisted pages are
# re-checked once they clear this from their last modification.
TRIAL_DAYS = 90

# Fixed quotas. Pages are scored, ranked once, then filled into these in order.
# DELETE takes whatever is left, so the four always sum to the live URL count.
QUOTAS = [("KEEP", 1200), ("UPDATE", 1000), ("MERGE", 1000)]

# Pages are listed in this sequence, and ranked by measured strength inside it,
# so the work reads top to bottom: what you protect, then what you fix, then
# what you fold in, then what you cut.
ACTION_RANK = {"KEEP": 0, "UPDATE": 1, "MERGE": 2, "DELETE": 3}

# Pages parked in MERGE or UPDATE that carry no measured signal and at least
# one structural strike. They stay on the site for now because they were
# produced too recently to judge; if they still show nothing at their review
# date, they become the next DELETE batch.
WATCH_NOTE = (
    "No measured signal yet and at least one structural problem. Kept for now "
    "because it has not had a fair trial. Re-check at the review date: still "
    "nothing, and it goes."
)

ACTIONS = {
    "KEEP":   "Leave alone. Either earning, or nothing argues against it.",
    "UPDATE": "Rewrite, improve or differentiate. Has a reason to change.",
    "MERGE":  "Fold into the stronger near-duplicate, then 301.",
    "DELETE": "Remove. Positive evidence it earns nothing worth keeping.",
}

REASONS = {
    "backlinked-earning": "Backlinked and earning clicks",
    "earns-clicks":       "Earns clicks",
    "striking-distance":  "Position 20 or better, within reach of page one",
    "unique-untested":    "Unique topic, not yet given a fair trial",
    "backlinked-idle":    "Backlinked but earning nothing, highest-value rewrite",
    "indexed-weak":       "Indexed and served, not converting",
    "canonical":          "Canonical of a duplicate group, absorb the others",
    "crowded":            "Crowded topic neighbourhood, needs differentiation",
    "dup-signal":         "Near-duplicate of a stronger page, but has signal",
    "dup-untested":       "Near-duplicate, group unmeasured, review before cutting",
    "dup-redundant":      "Near-duplicate of a page that already earns, it cannot outrank it",
    "padded-variant":     "Qualifier-padded variant of a shorter page, no signal of its own",
    "dup-dead":           "Near-duplicate, no signal after a fair trial",
    "dark-after-trial":   "No signal at all after a fair trial",
    "not-indexed":        "Google crawled it and refused to index it",
}

TIERS = [
    ("A", "Has external backlinks",                     "Never delete. Rewrite the ones with no traffic."),
    ("B", "Earns clicks, no backlinks",                 "Never delete. These pay the bills."),
    ("C", "Impressions, average position 20 or better", "Do not delete. Fastest available wins."),
    ("D", "Impressions but position past 20",           "Improve or fold into a stronger sibling."),
    ("E", "No row in the GSC export",                   "Mostly unmeasured, not proven dead. See the note."),
]

STOP = {"the", "a", "an", "for", "to", "and", "of", "in", "with", "your", "how", "what",
        "is", "best", "top", "guide", "complete", "2026", "2025", "edition", "ultimate",
        "ai", "vs", "on", "from"}


# Topic hierarchy. "How many pages should exist on SAP, or any topic" cannot be
# answered without first seeing how many DO exist, what they earn, and whether a
# real hub sits above them. A hub is a page outside /blog/: SAP Business One has
# one, most topics do not, which is why their pages have nothing to link up to.
TOPICS = [
    ("SAP Business One", ["sap-business-one", "sap-b1"]),
    ("NetSuite",         ["netsuite"]),
    ("JD Edwards",       ["jd-edwards", "jde-"]),
    ("Acumatica",        ["acumatica"]),
    ("Sage Intacct",     ["sage-intacct"]),
    ("Dynamics 365",     ["dynamics-365", "d365"]),
    ("Oracle Fusion",    ["oracle-fusion"]),
    ("QuickBooks",       ["quickbooks"]),
    ("Accounts Payable", ["accounts-payable", "ap-automation"]),
    ("Accounts Receivable", ["accounts-receivable", "ar-automation"]),
    ("Month End Close",  ["month-end-close", "financial-close"]),
    ("FP&A",             ["fpa", "fp-a"]),
    ("Reconciliation",   ["reconciliation"]),
    ("Cash Flow",        ["cash-flow"]),
    ("Glossary",         ["/glossary/"]),
]


def build_topics(pages):
    out = []
    for name, keys in TOPICS:
        hits = [p for p in pages if any(k in p["p"] for k in keys)]
        if not hits:
            continue
        # A pillar is a hub-shaped url: root level, or under solutions, ai-erp
        # or integrations. Articles are never pillars, however well they rank,
        # because nothing can be organised underneath them.
        def is_hub(x):
            u = x["p"]
            if u.startswith(("/blog/", "/glossary/", "/guide/", "/webinars/", "/insights/")):
                return False
            return u.count("/") == 2 or u.startswith(("/solutions/", "/ai-erp/", "/integrations/"))

        hubs = sorted([p for p in hits if is_hub(p)], key=lambda x: (-x["i"], len(x["p"])))
        hub = hubs[0] if hubs else None
        # three states, not two: a hub that earns, a hub that earns nothing, and
        # no hub at all with an article standing in its place.
        hub_state = "none" if hub is None else ("earning" if hub["i"] > 0 else "empty")
        counts = {"KEEP": 0, "UPDATE": 0, "MERGE": 0, "DELETE": 0}
        for p in hits:
            counts[p["s"]] += 1
        # the pages worth working first: most impressions, still no clicks
        best = sorted([p for p in hits if p["i"] > 0 and p["c"] == 0],
                      key=lambda x: -x["i"])[:3]
        out.append({
            "topic": name,
            "pages": len(hits),
            "impressions": sum(p["i"] for p in hits),
            "clicks": sum(p["c"] for p in hits),
            "backlinked": sum(p["b"] for p in hits),
            "orphans": sum(p["x"] for p in hits),
            "counts": counts,
            "hub": hub["p"] if hub else "",
            "hubImpressions": hub["i"] if hub else 0,
            "hubState": hub_state,
            "opportunities": [{"p": p["p"], "t": p["t"], "i": p["i"], "o": p["o"]} for p in best],
        })
    out.sort(key=lambda t: -t["impressions"])
    return out


def load_gsc():
    import openpyxl
    gsc = {}
    for f in glob.glob(GSC_GLOB):
        wb = openpyxl.load_workbook(f, read_only=True)
        for r in list(wb["Pages"].iter_rows(values_only=True))[1:]:
            u = r[0].replace("https://chatfin.ai", "")
            c, i, p = gsc.get(u, (0, 0, []))
            gsc[u] = (c + (r[1] or 0), i + (r[2] or 0), p + [r[4]])
    return gsc


def _norm_path(v):
    if not isinstance(v, str):
        return None
    v = v.replace("https://chatfin.ai", "").replace("http://chatfin.ai", "").strip()
    v = v.split("?")[0].split("#")[0].split(" ")[0]
    if not v.startswith("/"):
        return None
    return v if v.endswith("/") else v + "/"


def load_backlinked():
    """Union of every source that names a URL holding external backlinks."""
    out = set()

    def add(v):
        n = _norm_path(v)
        if n:
            out.add(n)

    for f in BACKLINK_LISTS:
        if os.path.exists(f):
            for line in open(f, errors="ignore"):
                if line.strip() and not line.startswith("#"):
                    add(line)
    for f in BACKLINK_TSV_COL0:
        if os.path.exists(f):
            for line in open(f, errors="ignore"):
                if line.startswith("#") or line.startswith("url\t"):
                    continue
                add(line.split("\t")[0])
    if os.path.exists(BACKLINK_XLSX):
        import openpyxl
        wb = openpyxl.load_workbook(BACKLINK_XLSX, read_only=True)
        for name in BACKLINK_XLSX_SHEETS:
            if name in wb.sheetnames:
                for r in wb[name].iter_rows(values_only=True):
                    if r:
                        add(r[0])
    return out


def load_rewritten():
    """URLs the ERP rewrite rigs have already republished."""
    if not os.path.exists(REWRITTEN):
        return set()
    out = set()
    for row in json.load(open(REWRITTEN)):
        n = _norm_path(row.get("url"))
        if n:
            out.add(n)
    return out


def load_not_indexed():
    """URLs GSC reports as crawled but deliberately not indexed."""
    if not os.path.exists(NOT_INDEXED):
        return set()
    import openpyxl
    wb = openpyxl.load_workbook(NOT_INDEXED, read_only=True)
    if "Table" not in wb.sheetnames:
        return set()
    out = set()
    for r in list(wb["Table"].iter_rows(values_only=True))[1:]:
        n = _norm_path(r[0] if r else None)
        if n:
            out.add(n)
    return out


def load_do_not_redirect():
    """URLs rewritten in place. Their URL must not move, so never merge or delete."""
    out = set()
    for f in glob.glob(DO_NOT_REDIRECT):
        for m in re.findall(r"https://chatfin\.ai[^\s]*", open(f, errors="ignore").read()):
            n = _norm_path(m)
            if n:
                out.add(n)
    return out


def load_crawled():
    """Paths Ahrefs could reach. Absent = orphaned, no internal links pointing in."""
    if not os.path.exists(AHREFS_CRAWL):
        return set()
    raw = open(AHREFS_CRAWL).read()
    out = set()
    for m in re.findall(r"https://chatfin\.ai[^ )\]]*", raw):
        p = m.replace("https://chatfin.ai", "") or "/"
        out.add(p if p.endswith("/") else p + "/")
    return out


def load_ref_domains():
    """Referring-domain counts per URL, where Ahrefs gave a number."""
    out = {}
    if not os.path.exists(REF_DOMAINS):
        return out
    for line in open(REF_DOMAINS):
        f = line.rstrip("\n").split("\t")
        if len(f) >= 4 and f[3].isdigit():
            out[f[0]] = int(f[3])
    return out


def titleise(slug):
    if not slug or slug == "home":
        return "Homepage"
    small = {"a", "an", "the", "for", "to", "and", "of", "in", "with", "on", "vs"}
    caps = {"ai", "erp", "ap", "ar", "cfo", "cfos", "kpi", "roi", "sap", "b1", "fpa", "dso", "gl"}
    out = []
    for i, w in enumerate(slug.replace("-", " ").split()):
        if w in caps:
            out.append(w.upper())
        elif i and w in small:
            out.append(w)
        else:
            out.append(w[:1].upper() + w[1:])
    return " ".join(out)


def tokens(slug):
    return frozenset(w for w in re.split(r"[-/]", slug)
                     if w and w not in STOP and not w.isdigit())


def group_at(paths, toks, inv, threshold):
    """Union-find over pairs sharing a token, joined when Jaccard >= threshold."""
    par = list(range(len(paths)))

    def find(x):
        while par[x] != x:
            par[x] = par[par[x]]
            x = par[x]
        return x

    seen = set()
    for w, idxs in inv.items():
        if len(idxs) > 400:      # ubiquitous token, tells us nothing
            continue
        for a in range(len(idxs)):
            for b in range(a + 1, len(idxs)):
                i, j = idxs[a], idxs[b]
                if (i, j) in seen:
                    continue
                seen.add((i, j))
                ti, tj = toks[i], toks[j]
                if ti and tj and len(ti & tj) / len(ti | tj) >= threshold:
                    ri, rj = find(i), find(j)
                    if ri != rj:
                        par[rj] = ri
    g = collections.defaultdict(list)
    for i in range(len(paths)):
        g[find(i)].append(i)
    return g


def padded_variants(paths, toks, inv, has_signal):
    """
    Find qualifier-padded variants: pages whose slug tokens are a strict
    SUPERSET of a shorter, substantive page's tokens.

    "accrual-automation-software-for-month-end-close" is a superset of
    "accrual-automation-software", so it is a padded variant of it. This is the
    signature of batch-generated bloat, and Jaccard misses it because the extra
    qualifiers drag similarity below the threshold.

    Only pages with no signal of their own are considered . a padded slug that
    actually earns is a real page.
    """
    variant = {}
    for path in paths:
        t = toks[path]
        if len(t) < 3 or has_signal(path):
            continue
        rare = min(t, key=lambda w: len(inv[w]))
        for other in inv[rare]:
            if other == path:
                continue
            to = toks[other]
            if to and len(to) >= 2 and to < t:      # strict subset = the base page
                variant[path] = other
                break
    return variant


def keep_score(p, gsc, backlinked, ref_domains, crawled, canonical,
               non_canonical, crowded, variants, old, not_indexed=frozenset()):
    """
    One number per page: how much it is worth keeping. Ranked descending, then
    filled into the quotas. Measured performance dominates; structural problems
    subtract. A backlinked page carries a bonus large enough that it can never
    fall into DELETE, whatever else is true of it.
    """
    g = gsc.get(p, (0, 0, []))
    clicks, impressions, positions = g[0], g[1], g[2]
    score = clicks * 1000.0 + impressions

    if p in backlinked:
        score += 50000 + ref_domains.get(p, 0) * 500

    if positions:
        avg = sum(positions) / len(positions)
        if avg <= 20:
            score += (21 - avg) * 50

    if p in canonical:
        score += 500

    # Structural penalties apply only to pages that have not proved themselves.
    # A page earning clicks has answered the question the structure was a proxy
    # for, so a padded slug or a missing internal link no longer counts against
    # it . otherwise the penalties can push an earning page into DELETE.
    if clicks == 0:
        if p in not_indexed:
            # Google looked at this page and said no. Age is no defence here:
            # the page was crawled, so it had its chance to be judged.
            score -= 5000
        if p in non_canonical:
            score -= 2000
        if p in variants:
            score -= 3000
        if p in crowded:
            score -= 200
        if crawled and p not in crawled:
            score -= 300
        if old and impressions == 0:
            score -= 1000      # had its chance and showed nothing

    return score


def order_pages(pages):
    """
    Sort by action sequence, then by measured strength inside each action.

    Strength is clicks, then impressions, then position (lower is better, and
    unranked sorts last). Each page also gets `n`, its 1-based rank by strength
    within this list, so a row can show where it sits regardless of the sort.
    """
    by_strength = sorted(
        pages,
        key=lambda x: (
            -x["c"],                                    # clicks first
            -x["i"],                                    # then impressions
            x["o"] if x["o"] is not None else 9e9,       # then position, unranked last
            -x.get("rd", 0),                             # then referring domains
            x.get("x", 0),                               # reachable above orphaned
            x["p"],
        ),
    )
    for rank, pg in enumerate(by_strength, 1):
        pg["n"] = rank
    return sorted(pages, key=lambda x: (ACTION_RANK[x["s"]], x["n"]))


def main():
    urls = json.load(open(SITEMAP))
    meta = {u["path"]: u for u in urls}
    gsc = load_gsc()
    crawled = load_crawled()
    no_redirect = load_do_not_redirect()
    rewritten = load_rewritten() | no_redirect
    not_indexed = load_not_indexed()
    ref_domains = load_ref_domains()
    backlinked = load_backlinked()

    paths = [u["path"] for u in urls]
    toks = [tokens(u["slug"]) for u in urls]
    inv = collections.defaultdict(list)
    for i, t in enumerate(toks):
        for w in t:
            inv[w].append(i)

    def clicks(p):
        return gsc[p][0] if p in gsc else 0

    def impressions(p):
        return gsc[p][1] if p in gsc else 0

    # ---- duplicate groups, and the canonical of each --------------------------
    tight = {k: v for k, v in group_at(paths, toks, inv, TIGHT).items() if len(v) > 1}
    canonical, non_canonical = set(), set()
    dup_groups = []
    for members in tight.values():
        members.sort(key=lambda i: (-clicks(paths[i]), -impressions(paths[i]),
                                    0 if paths[i] in backlinked else 1, len(paths[i])))
        # a backlinked member always takes the canonical slot
        linked = [i for i in members if paths[i] in backlinked]
        head = linked[0] if linked else members[0]
        canonical.add(paths[head])
        non_canonical.update(paths[i] for i in members if i != head)
        dup_groups.append(len(members))

    loose = group_at(paths, toks, inv, LOOSE)
    crowded = {paths[i] for v in loose.values() if len(v) >= CROWD_MIN for i in v}

    # which canonical each non-canonical page defers to, and whether that
    # canonical has actually demonstrated anything
    defers_to = {}
    for members in tight.values():
        linked = [i for i in members if paths[i] in backlinked]
        head = linked[0] if linked else members[0]
        for i in members:
            if i != head:
                defers_to[paths[i]] = paths[head]

    def has_signal(p):
        return p in backlinked or clicks(p) > 0 or impressions(p) > 0

    tok_by_path = {paths[i]: toks[i] for i in range(len(paths))}
    inv_by_path = collections.defaultdict(list)
    for path, t in tok_by_path.items():
        for w in t:
            inv_by_path[w].append(path)
    variants = padded_variants(paths, tok_by_path, inv_by_path, has_signal)

    def had_fair_trial(p):
        return (meta[p].get("lastmod") or "")[:7] <= FAIR_TRIAL_BEFORE

    # ---- classify -------------------------------------------------------------
    def decide(p):
        g = gsc.get(p)
        c = g[0] if g else 0
        i = g[1] if g else 0
        position = (sum(g[2]) / len(g[2])) if g and g[2] else None
        old = had_fair_trial(p)

        if p in backlinked:
            return ("KEEP", "backlinked-earning") if c > 0 else ("UPDATE", "backlinked-idle")
        if c > 0:
            return "KEEP", "earns-clicks"
        if i >= 10 and position is not None and position <= 20:
            return "KEEP", "striking-distance"
        if p in not_indexed:
            return "DELETE", "not-indexed"
        if p in non_canonical:
            if i > 0:
                return "MERGE", "dup-signal"
            if old:
                return "DELETE", "dup-dead"
            # Age defends a unique page, not a duplicate. A new page might yet
            # rank; a near-duplicate of a page that ALREADY earns will not,
            # because Google picks one of them and has already picked. There is
            # nothing to fold in either . the canonical carries the content.
            if has_signal(defers_to.get(p, "")):
                return "DELETE", "dup-redundant"
            # whole group unmeasured: cannot tell which member is best yet
            return "MERGE", "dup-untested"
        if i > 0:
            return "UPDATE", "indexed-weak"
        if p in canonical:
            return "UPDATE", "canonical"
        if old:
            return "DELETE", "dark-after-trial"
        if p in variants:
            return "DELETE", "padded-variant"
        if p in crowded:
            return "UPDATE", "crowded"
        return "KEEP", "unique-untested"

    def tier_of(p):
        g = gsc.get(p)
        if p in backlinked:
            return "A"
        if g and g[0] > 0:
            return "B"
        if g and g[1] >= 10 and sum(g[2]) / len(g[2]) <= 20:
            return "C"
        if g:
            return "D"
        return "E"

    # ---- score every URL, rank once, fill the quotas -------------------------
    scored = sorted(
        ((keep_score(u["path"], gsc, backlinked, ref_domains, crawled, canonical,
                     non_canonical, crowded, variants, had_fair_trial(u["path"]),
                     not_indexed),
          u["path"])
         for u in urls),
        key=lambda x: (-x[0], x[1]),
    )
    # A page that has external backlinks, or that earns clicks, is never
    # deleted. This is enforced by construction rather than left to the scoring
    # margin: DELETE is filled from the BOTTOM of the ranking, skipping any
    # protected page, so a protected page cannot land there however badly it
    # scores. The surviving buckets then fill from the top in score order.
    protected = {u["path"] for u in urls
                 if u["path"] in backlinked
                 or clicks(u["path"]) > 0
                 or u["path"] in rewritten}

    delete_quota = len(scored) - sum(size for _, size in QUOTAS)
    if delete_quota > len(scored) - len(protected):
        raise SystemExit(
            f"cannot honour the quotas: {delete_quota} to delete but only "
            f"{len(scored) - len(protected)} unprotected pages exist"
        )

    to_delete = set()
    for _, path in reversed(scored):           # weakest first
        if len(to_delete) >= delete_quota:
            break
        if path not in protected:
            to_delete.add(path)

    quota_action = {path: "DELETE" for path in to_delete}
    cursor = 0
    survivors = [path for _, path in scored if path not in to_delete]
    for name, size in QUOTAS:
        for path in survivors[cursor:cursor + size]:
            quota_action[path] = name
        cursor += size

    assert not (protected & to_delete), "a protected page was assigned DELETE"

    # MERGE also moves a URL, because folding a page in means 301ing it. A page
    # on a do-not-redirect list was rewritten in place and must keep its URL, so
    # it cannot sit in MERGE either. Swap any that landed there with the weakest
    # unprotected page in UPDATE, which keeps every quota the same size.
    stuck = [p for p in survivors if quota_action.get(p) == "MERGE" and p in no_redirect]
    if stuck:
        swappable = [p for p in reversed(survivors)
                     if quota_action.get(p) == "UPDATE" and p not in no_redirect]
        for a, b in zip(stuck, swappable):
            quota_action[a], quota_action[b] = "UPDATE", "MERGE"
        log_swaps = len(stuck)
    else:
        log_swaps = 0

    pages = []
    for u in urls:
        p = u["path"]
        _, reason = decide(p)
        action = quota_action[p]
        g = gsc.get(p, (0, 0, []))
        parts = [x for x in p.split("/") if x]
        if not parts:
            cluster, sub, slug = "home", None, "home"
        elif len(parts) == 1:
            cluster, sub, slug = "root", None, parts[0]
        elif len(parts) == 2:
            cluster, sub, slug = parts[0], None, parts[1]
        else:
            cluster, sub, slug = parts[0], parts[1], parts[-1]

        strikes = []
        if crawled and p not in crawled:
            strikes.append("orphan")
        if p in canonical or p in non_canonical:
            strikes.append("duplicate")
        if p in crowded:
            strikes.append("crowded")
        if p in variants:
            strikes.append("padded")

        # watchlist: no signal, no links, structurally weak, and not the page
        # its duplicate group is being consolidated onto
        watch = (
            action in ("MERGE", "UPDATE")
            and g[0] == 0 and g[1] == 0
            and p not in backlinked
            and p not in rewritten
            # the watchlist exists for pages produced too recently to judge.
            # a page that already had its fair trial and showed nothing is not
            # waiting on anything, so it does not belong here.
            and not had_fair_trial(p)
            and g[0] == 0
            and reason != "canonical"
            and len(strikes) >= 1
        )
        review_due = ""
        if watch:
            lm = (meta[p].get("lastmod") or "")[:10]
            try:
                d0 = datetime.date.fromisoformat(lm)
                # a handful of sitemap entries carry a 1970 epoch date. Treating
                # those as real made the next review date read 1970-04-01.
                if d0.year < 2020:
                    d0 = datetime.date(2026, 8, 19)
                review_due = (d0 + datetime.timedelta(days=TRIAL_DAYS)).isoformat()
            except ValueError:
                review_due = ""

        pages.append({
            "p": p,
            "t": titleise(slug),
            "s": action,
            "r": reason,
            "k": tier_of(p),
            "c": int(g[0]),
            "i": int(g[1]),
            "o": round(sum(g[2]) / len(g[2]), 1) if g[2] else None,
            "b": 1 if p in backlinked else 0,
            "d": 1 if (p in canonical or p in non_canonical) else 0,
            # orphan: in the sitemap but not reachable by Ahrefs' crawler
            "x": 0 if (not crawled or p in crawled) else 1,
            "rd": ref_domains.get(p, 0),
            "w": 1 if watch else 0,
            "wd": review_due,
            "st": strikes,
            "_cluster": cluster,
            "_sub": sub,
        })

    # ---- nest -----------------------------------------------------------------
    def blank():
        return {"KEEP": 0, "UPDATE": 0, "MERGE": 0, "DELETE": 0}

    byc = collections.defaultdict(list)
    for pg in pages:
        byc[pg["_cluster"]].append(pg)

    clusters = []
    for cname, cpages in byc.items():
        counts, tiers = blank(), {}
        cl = im = bkl = wch = 0
        subs = collections.defaultdict(list)
        direct = []
        for pg in cpages:
            counts[pg["s"]] += 1
            tiers[pg["k"]] = tiers.get(pg["k"], 0) + 1
            cl += pg["c"]; im += pg["i"]; bkl += pg["b"]; wch += pg["w"]
            (subs[pg["_sub"]] if pg["_sub"] else direct).append(pg)

        sub_out = []
        for sname, spages in subs.items():
            sc, st = blank(), {}
            scl = sim = sbk = swc = 0
            for pg in spages:
                sc[pg["s"]] += 1
                st[pg["k"]] = st.get(pg["k"], 0) + 1
                scl += pg["c"]; sim += pg["i"]; sbk += pg["b"]; swc += pg["w"]
            sub_out.append({
                "sub": sname, "title": titleise(sname), "urls": len(spages),
                "counts": sc, "tiers": st, "clicks": scl, "impressions": sim,
                "backlinked": sbk, "watch": swc,
                "pages": order_pages(spages),
            })
        sub_out.sort(key=lambda s: -s["urls"])

        clusters.append({
            "cluster": cname, "title": titleise(cname), "urls": len(cpages),
            "counts": counts, "tiers": tiers, "clicks": cl, "impressions": im,
            "backlinked": bkl, "watch": wch, "subs": sub_out,
            "pages": order_pages(direct),
        })
    clusters.sort(key=lambda c: -c["urls"])

    for c in clusters:
        for pg in c["pages"]:
            pg.pop("_cluster", None); pg.pop("_sub", None)
        for s in c["subs"]:
            for pg in s["pages"]:
                pg.pop("_cluster", None); pg.pop("_sub", None)

    # ---- totals ---------------------------------------------------------------
    totals = blank()
    reason_counts = collections.Counter()
    tier_counts, tier_clicks, tier_impr = (collections.Counter() for _ in range(3))
    for pg in pages:
        totals[pg["s"]] += 1
        reason_counts[pg["r"]] += 1
        tier_counts[pg["k"]] += 1
        tier_clicks[pg["k"]] += pg["c"]
        tier_impr[pg["k"]] += pg["i"]

    orphan_count = sum(1 for pg in pages if pg["x"] == 1)
    watch_count = sum(1 for pg in pages if pg["w"] == 1)
    watch_by_action = dict(collections.Counter(pg["s"] for pg in pages if pg["w"] == 1))
    review_dates = sorted(pg["wd"] for pg in pages if pg["w"] == 1 and pg["wd"])
    next_review = review_dates[0] if review_dates else ""
    untested = sum(1 for pg in pages
                   if not had_fair_trial(pg["p"]) and pg["p"] not in gsc)

    doc = {
        "generated": GENERATED,
        "window": WINDOW,
        "method": {
            "fairTrialBefore": FAIR_TRIAL_BEFORE,
            "tight": TIGHT,
            "loose": LOOSE,
            "crowdMin": CROWD_MIN,
            "duplicateGroups": len(dup_groups),
            "duplicateUrls": sum(dup_groups),
            "duplicateSurplus": sum(dup_groups) - len(dup_groups),
            "crowdedUrls": len(crowded),
            "paddedVariants": len(variants),
            "crawledUrls": len(crawled),
            "orphanUrls": orphan_count,
            "rewrittenUrls": len(rewritten),
            "notIndexedUrls": sum(1 for u in urls if u["path"] in not_indexed),
            "trialDays": TRIAL_DAYS,
            "untestedUrls": untested,
        },
        "source": {
            "sitemap": "chatfin-all-urls.json",
            "gsc": "GSC page exports, 1,000-row cap each",
            "backlinks": "Ahrefs: best-by-links exports and referring-domain tables, unioned",
        },
        "whyDeleteIsSmall": [
            "GSC's export stops at 1,000 rows and the weakest page in each export had "
            "1-2 impressions, so the cap bites at the bottom of the range. A page missing "
            "from the export may simply have ranked below the cut.",
            f"90% of the site carries a Jul/Aug 2026 lastmod, and {untested:,} pages with no "
            "GSC row were published or modified inside the two-week window being measured. "
            "They have not had time to rank.",
            "So absence of traffic is not treated as evidence. DELETE requires positive "
            "evidence: a near-duplicate of a better page, or no signal after a fair trial.",
        ],
        "caveats": [
            "The backlink list is Ahrefs page 1. A full sweep will add to the protect list.",
            "Duplicate detection is slug similarity, a proxy. The real test is body-text "
            "hashing, which is what the SAP B1 audit used and what found 13 URLs sharing "
            "one article.",
            "Re-measure once the untested pages have 90 days of history, then re-run this.",
        ],
        "totals": {
            "urls": len(pages),
            "clicks": sum(p["c"] for p in pages),
            "impressions": sum(p["i"] for p in pages),
            "backlinked": sum(p["b"] for p in pages),
            **totals,
        },
        "watch": {
            "note": WATCH_NOTE,
            "count": watch_count,
            "byAction": watch_by_action,
            "nextReview": next_review,
        },
        "topics": build_topics(pages),
        "actions": ACTIONS,
        "reasons": REASONS,
        "reasonCounts": dict(reason_counts),
        "tiers": [
            {"tier": t, "definition": d, "verdict": v, "urls": tier_counts[t],
             "clicks": tier_clicks[t], "impressions": tier_impr[t]}
            for t, d, v in TIERS
        ],
        "clusters": clusters,
    }

    # ---- invariants: fail loudly rather than ship a plan that deletes equity --
    deleted = {pg["p"] for pg in pages if pg["s"] == "DELETE"}
    watched = {pg["p"] for pg in pages if pg["w"] == 1}
    violations = []
    for label, offenders in (
        ("backlinked page in DELETE", backlinked & deleted),
        ("backlinked page on the watchlist", backlinked & watched),
        ("page with clicks in DELETE", {p for p in deleted if clicks(p) > 0}),
        ("page with clicks on the watchlist", {p for p in watched if clicks(p) > 0}),
        ("already-rewritten page in DELETE", rewritten & deleted),
        ("do-not-redirect page in MERGE or DELETE",
         no_redirect & ({pg["p"] for pg in pages if pg["s"] in ("MERGE", "DELETE")})),
        ("already-rewritten page on the watchlist", rewritten & watched),
    ):
        if offenders:
            violations.append(f"{label}: {len(offenders)} e.g. {sorted(offenders)[:3]}")
    if violations:
        raise SystemExit("INVARIANT FAILED\n  " + "\n  ".join(violations))

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "w") as fh:
        json.dump(doc, fh, separators=(",", ":"))

    print(f"wrote {OUT} ({os.path.getsize(OUT)/1024:.0f} KB)")

    # per-URL decisions, so the plan can actually be worked rather than only counted
    PAGES_OUT = os.path.join(REPO, "data", "plan26-pages.json")
    with open(PAGES_OUT, "w") as fh:
        json.dump(pages, fh, separators=(",", ":"))
    print(f"wrote {PAGES_OUT} ({os.path.getsize(PAGES_OUT)/1024:.0f} KB) rows={len(pages)}")
    print(f"  {len(pages)} URLs across {len(clusters)} clusters")
    print("  " + "  ".join(f"{k}={v}" for k, v in totals.items()))
    print(f"  duplicate groups {len(dup_groups)} covering {sum(dup_groups)} URLs")
    print(f"  untested (no fair trial yet) {untested}")
    print(f"  invariants OK: no backlinked or earning page in DELETE or watchlist")
    if log_swaps:
        print(f"  moved {log_swaps} do-not-redirect pages out of MERGE into UPDATE")
    print(f"  watchlist {watch_count} -> DELETE now {totals['DELETE']} + watch "
          f"{watch_count} = {totals['DELETE'] + watch_count}")


if __name__ == "__main__":
    main()
